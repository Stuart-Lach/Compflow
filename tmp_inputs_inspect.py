from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\Users\adria\Compflow\compliance-engine\data\sa_payroll_workbook.xlsx")
wb = load_workbook(wb_path, data_only=True)
ws = wb['Inputs']
lines = []
lines.append(f"max_row={ws.max_row} max_col={ws.max_column}")
for r in range(1, 10):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    lines.append(f"row{r}: {row}")
Path(r"C:\Users\adria\Compflow\tmp_inputs_inspect.txt").write_text("\n".join(lines), encoding="utf-8")
