from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\Users\adria\Compflow\compliance-engine\data\sa_payroll_workbook.xlsx")
print('exists', wb_path.exists())
wb = load_workbook(wb_path, data_only=True)
print(wb.sheetnames)
ws = wb['Outputs']
print('max_row', ws.max_row, 'max_col', ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8, values_only=True):
    print(row)
PY'@
$script | Set-Content -Encoding UTF8 C:\Users\adria\Compflow\tmp_outputs_inspect.py
python C:\Users\adria\Compflow\tmp_outputs_inspect.py
python C:\Users\adria\Compflow\tmp_outputs_inspect.py
$script = @'
from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\Users\adria\Compflow\compliance-engine\data\sa_payroll_workbook.xlsx")
print('exists', wb_path.exists())
wb = load_workbook(wb_path, data_only=True)
print(wb.sheetnames)
ws = wb['Outputs']
print('max_row', ws.max_row, 'max_col', ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=8, values_only=True):
    print(row)
