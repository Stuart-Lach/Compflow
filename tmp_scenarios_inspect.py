from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\Users\adria\Compflow\compliance-engine\data\sa_payroll_workbook.xlsx")
wb = load_workbook(wb_path, data_only=True)
ws = wb['Scenarios']
lines = []
for r in range(1, 6):
    row = [ws.cell(r, c).value for c in range(1, 7)]
    lines.append(f"row{r}: {row}")
Path(r"C:\Users\adria\Compflow\tmp_scenarios_inspect.txt").write_text("\n".join(lines), encoding="utf-8")
