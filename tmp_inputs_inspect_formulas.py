from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\Users\adria\Compflow\compliance-engine\data\sa_payroll_workbook.xlsx")
wb = load_workbook(wb_path, data_only=False)
ws = wb['Inputs']
lines = []
for r in range(1, 6):
    row = []
    for c in range(1, 7):
        cell = ws.cell(r, c)
        row.append(cell.value)
    lines.append(f"row{r}: {row}")
Path(r"C:\Users\adria\Compflow\tmp_inputs_inspect_formulas.txt").write_text("\n".join(lines), encoding="utf-8")
