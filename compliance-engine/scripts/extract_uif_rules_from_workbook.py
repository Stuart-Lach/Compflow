"""
Extract UIF rules from the workbook Ruleset sheet.

Writes JSON to data/samples/uif_rules_from_workbook.json.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


def extract_uif_rules(workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Ruleset v1.1 (2025-26)"]

    # UIF parameters are listed in A15:B17 per workbook layout
    ceiling = ws.cell(row=15, column=2).value
    employee_rate = ws.cell(row=16, column=2).value
    employer_rate = ws.cell(row=17, column=2).value

    return {
        "uif_monthly_ceiling": ceiling,
        "uif_employee_rate": employee_rate,
        "uif_employer_rate": employer_rate,
        "source": {
            "workbook": str(workbook_path),
            "sheet": "Ruleset v1.1 (2025-26)",
            "range": "A15:B17",
        },
    }


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    workbook_path = repo_root / "data" / "sa_payroll_workbook.xlsx"
    output_path = repo_root / "data" / "samples" / "uif_rules_from_workbook.json"

    data = extract_uif_rules(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
