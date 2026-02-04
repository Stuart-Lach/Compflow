"""
Extract PAYE brackets, rebates, and thresholds from the workbook Ruleset sheet.

Writes JSON to data/samples/paye_rules_from_workbook.json.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class Bracket:
    from_amount: int
    to_amount: Optional[int]
    base_tax: int
    marginal_rate: float


def _find_first_row(ws, label: str) -> int:
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=8):
        for cell in row:
            if cell.value == label:
                return cell.row
    raise ValueError(f"Label not found: {label}")


def extract_ruleset(workbook_path: Path) -> Dict[str, object]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Ruleset v1.1 (2025-26)"]

    brackets_start = _find_first_row(ws, "PAYE Tax Brackets (Annual)") + 2
    brackets: List[Bracket] = []

    row = brackets_start
    while True:
        from_val = ws.cell(row=row, column=1).value
        to_val = ws.cell(row=row, column=2).value
        base_tax = ws.cell(row=row, column=3).value
        rate = ws.cell(row=row, column=4).value
        if from_val is None:
            break
        brackets.append(
            Bracket(
                from_amount=int(from_val),
                to_amount=int(to_val) if to_val is not None else None,
                base_tax=int(base_tax),
                marginal_rate=float(rate),
            )
        )
        row += 1

    rebates_start = _find_first_row(ws, "Tax Rebates (Annual)") + 2
    rebates: Dict[str, int] = {}

    row = rebates_start
    while True:
        category = ws.cell(row=row, column=6).value
        rebate = ws.cell(row=row, column=7).value
        if category is None:
            break
        rebates[str(category).strip().lower()] = int(rebate)
        row += 1

    thresholds_start = _find_first_row(ws, "Tax Thresholds (Annual)") + 2
    thresholds: Dict[str, int] = {}

    row = thresholds_start
    while True:
        category = ws.cell(row=row, column=6).value
        threshold = ws.cell(row=row, column=7).value
        if category is None:
            break
        thresholds[str(category).strip().lower()] = int(threshold)
        row += 1

    return {
        "brackets": [asdict(b) for b in brackets],
        "rebates": rebates,
        "thresholds": thresholds,
        "source": {
            "workbook": str(workbook_path),
            "sheet": "Ruleset v1.1 (2025-26)",
        },
    }


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    workbook_path = repo_root / "data" / "sa_payroll_workbook.xlsx"
    output_path = repo_root / "data" / "samples" / "paye_rules_from_workbook.json"

    data = extract_ruleset(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
