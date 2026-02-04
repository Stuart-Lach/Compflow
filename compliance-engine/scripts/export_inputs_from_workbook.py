"""
Export Inputs sheet from workbook to CSV (openpyxl).

Writes CSV to data/samples/inputs_from_workbook.csv.
"""

from __future__ import annotations

import csv
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


def _normalize_header(value: str) -> str:
    return (
        value.strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("(", "")
        .replace(")", "")
    )


_FORMULA_REF = re.compile(r"^=Scenarios!([A-Z]+)(\d+)$")


def _resolve_value(formula_or_value, scenarios_ws):
    if isinstance(formula_or_value, str):
        match = _FORMULA_REF.match(formula_or_value.strip())
        if match:
            col, row = match.group(1), int(match.group(2))
            return scenarios_ws[f"{col}{row}"].value
    return formula_or_value


def export_inputs(workbook_path: Path, output_path: Path) -> None:
    wb_formulas = load_workbook(workbook_path, data_only=False)
    wb_values = load_workbook(workbook_path, data_only=True)
    ws = wb_formulas["Inputs"]
    scenarios_ws = wb_values["Scenarios"]

    header_row = None
    for row_idx in range(1, min(ws.max_row, 50) + 1):
        row_values = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if any(str(v).strip().lower() in {"employee_id", "employee id"} for v in row_values if v is not None):
            header_row = row_idx
            break

    if header_row is None:
        raise ValueError("Inputs header row not found (expected 'employee_id').")

    raw_headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers = [
        _normalize_header(str(h)) if h is not None and str(h).strip() else ""
        for h in raw_headers
    ]

    if not any(headers):
        raise ValueError("Inputs header row is empty.")

    header_map = {name: idx for idx, name in enumerate(headers) if name}

    output_path.parent.mkdir(parents=True, exist_ok=True)

    canonical_headers = [
        "payroll_run_id",
        "company_id",
        "pay_date",
        "tax_year",
        "payroll_frequency",
        "employee_id",
        "employment_type",
        "basic_salary",
        "overtime_pay",
        "bonus_commission",
        "allowances_taxable",
        "other_post_tax_deductions",
    ]

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=canonical_headers)
        writer.writeheader()

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_values = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in row_values):
                break

            def get_col(name: str):
                idx = header_map.get(name)
                if idx is None:
                    return None
                return _resolve_value(row_values[idx], scenarios_ws)

            employee_id = get_col("employee_id")
            if not employee_id:
                break

            record = {
                "payroll_run_id": "RUN_WORKBOOK",
                "company_id": "COMP_WORKBOOK",
                "pay_date": date(2025, 4, 1).isoformat(),
                "tax_year": "2025_26",
                "payroll_frequency": "monthly",
                "employee_id": employee_id,
                "employment_type": "employee",
                "basic_salary": get_col("basic_salary_monthly") or get_col("basic_salary"),
                "overtime_pay": get_col("overtime_monthly") or get_col("overtime"),
                "bonus_commission": get_col("bonus_monthly") or get_col("bonus"),
                "allowances_taxable": get_col("allowances_monthly") or get_col("allowances"),
                "other_post_tax_deductions": get_col("other_deductions_monthly") or get_col("other_deductions"),
            }

            writer.writerow({k: "" if v is None else v for k, v in record.items()})


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    workbook_path = repo_root / "data" / "sa_payroll_workbook.xlsx"
    output_path = repo_root / "data" / "samples" / "inputs_from_workbook.csv"

    export_inputs(workbook_path, output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
