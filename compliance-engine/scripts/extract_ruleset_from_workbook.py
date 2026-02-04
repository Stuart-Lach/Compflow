"""
Extract ruleset data from the workbook (openpyxl).

Writes JSON to data/samples/ruleset_from_workbook.json.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class Bracket:
    from_amount: int
    to_amount: Optional[int]
    base_tax: int
    marginal_rate: float


def _normalize_key(value: str) -> str:
    return (
        value.strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("(", "")
        .replace(")", "")
    )


def extract_ruleset(workbook_path: Path) -> Dict[str, object]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Ruleset v1.1 (2025-26)"]

    # PAYE brackets: A4:D until blank in column A
    brackets: List[Bracket] = []
    row = 4
    while True:
        from_val = ws.cell(row=row, column=1).value
        to_val = ws.cell(row=row, column=2).value
        base_tax = ws.cell(row=row, column=3).value
        rate = ws.cell(row=row, column=4).value
        if from_val is None:
            break
        try:
            from_amount = int(from_val)
            to_amount = int(to_val) if to_val is not None else None
            base_amount = int(base_tax)
            rate_amount = float(rate)
        except (TypeError, ValueError):
            row += 1
            continue
        brackets.append(
            Bracket(
                from_amount=from_amount,
                to_amount=to_amount,
                base_tax=base_amount,
                marginal_rate=rate_amount,
            )
        )
        row += 1

    # Rebates: F4:G until blank in column F
    rebates: Dict[str, int] = {}
    row = 4
    while True:
        category = ws.cell(row=row, column=6).value
        rebate = ws.cell(row=row, column=7).value
        if category is None:
            break
        try:
            rebates[_normalize_key(str(category))] = int(rebate)
        except (TypeError, ValueError):
            pass
        row += 1

    # Thresholds: F10:G until blank in column F
    thresholds: Dict[str, int] = {}
    row = 10
    while True:
        category = ws.cell(row=row, column=6).value
        threshold = ws.cell(row=row, column=7).value
        if category is None:
            break
        try:
            thresholds[_normalize_key(str(category))] = int(threshold)
        except (TypeError, ValueError):
            pass
        row += 1

    # UIF params: A15:B17
    uif_params: Dict[str, float] = {}
    for row in range(15, 18):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if key is None or value is None:
            continue
        try:
            uif_params[_normalize_key(str(key))] = float(value)
        except (TypeError, ValueError):
            pass

    # SDL params: A18:B19
    sdl_params: Dict[str, float] = {}
    for row in range(18, 20):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if key is None or value is None:
            continue
        try:
            sdl_params[_normalize_key(str(key))] = float(value)
        except (TypeError, ValueError):
            pass

    return {
        "brackets": [asdict(b) for b in brackets],
        "rebates": rebates,
        "thresholds": thresholds,
        "uif": uif_params,
        "sdl": sdl_params,
        "source": {
            "workbook": str(workbook_path),
            "sheet": "Ruleset v1.1 (2025-26)",
        },
    }


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    workbook_path = repo_root / "data" / "sa_payroll_workbook.xlsx"
    output_path = repo_root / "data" / "samples" / "ruleset_from_workbook.json"

    data = extract_ruleset(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
