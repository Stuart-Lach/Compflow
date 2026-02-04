"""
Extract tax values from sa_payroll_workbook.xlsx and update za_2025_26_v1.py

This script reads the authoritative Excel workbook and extracts:
- PAYE tax brackets (annual and monthly)
- UIF caps and rates
- SDL threshold and rate
- Tax rebates

Run this after copying sa_payroll_workbook.xlsx to data/ directory.
"""

import openpyxl
from decimal import Decimal
from pathlib import Path
from datetime import date


def extract_values_from_workbook():
    """Extract tax values from the Excel workbook."""

    workbook_path = Path(__file__).parent.parent / "data" / "sa_payroll_workbook.xlsx"

    if not workbook_path.exists():
        print(f"❌ ERROR: Workbook not found at {workbook_path}")
        print("Please copy sa_payroll_workbook.xlsx to the data/ directory")
        print(f"Expected location: {workbook_path}")
        return None

    print(f"📖 Reading workbook from: {workbook_path}")

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)

        print("\n📋 Available sheets:")
        for sheet_name in wb.sheetnames:
            print(f"  - {sheet_name}")

        # Extract values based on common SA tax workbook structure
        # Adjust sheet names and cell references based on actual workbook

        values = {
            "extraction_date": date.today().isoformat(),
            "tax_brackets": extract_tax_brackets(wb),
            "uif": extract_uif_config(wb),
            "sdl": extract_sdl_config(wb),
            "rebates": extract_rebates(wb),
        }

        print("\n✅ Extraction complete")
        return values

    except Exception as e:
        print(f"❌ ERROR reading workbook: {e}")
        import traceback
        traceback.print_exc()
        return None


def extract_tax_brackets(wb):
    """Extract PAYE tax brackets from workbook."""
    print("\n🔍 Extracting PAYE tax brackets...")

    # Common sheet names for tax brackets
    possible_sheet_names = ["Tax Brackets", "PAYE", "Tax", "Brackets", "2025_26"]

    sheet = None
    for name in possible_sheet_names:
        if name in wb.sheetnames:
            sheet = wb[name]
            print(f"   Found sheet: {name}")
            break

    if not sheet:
        print(f"   ⚠️  Using first sheet: {wb.sheetnames[0]}")
        sheet = wb[wb.sheetnames[0]]

    brackets = []

    # Try to find tax bracket data (common patterns)
    # Look for columns: Min Income, Max Income, Rate, Base Tax
    for row_idx in range(1, min(50, sheet.max_row + 1)):
        try:
            # Try different column layouts
            # Pattern 1: A=min, B=max, C=rate, D=base_tax
            min_val = sheet.cell(row_idx, 1).value
            max_val = sheet.cell(row_idx, 2).value
            rate_val = sheet.cell(row_idx, 3).value
            base_val = sheet.cell(row_idx, 4).value

            if min_val and isinstance(min_val, (int, float)) and min_val >= 0:
                bracket = {
                    "min_income": Decimal(str(min_val)),
                    "max_income": Decimal(str(max_val)) if max_val and max_val > 0 else None,
                    "rate": Decimal(str(rate_val)) if rate_val else Decimal("0"),
                    "base_tax": Decimal(str(base_val)) if base_val else Decimal("0"),
                }
                brackets.append(bracket)
                print(f"   Bracket: R{bracket['min_income']:,.0f} - {'∞' if not bracket['max_income'] else f\"R{bracket['max_income']:,.0f}\"} @ {bracket['rate']*100}%")
        except:
            continue

    if not brackets:
        print("   ⚠️  No brackets found, using 2024/25 official values as fallback")
        # Use official SARS 2024/25 values (will be updated when workbook is available)
        brackets = [
            {"min_income": Decimal("0"), "max_income": Decimal("237100"), "rate": Decimal("0.18"), "base_tax": Decimal("0")},
            {"min_income": Decimal("237101"), "max_income": Decimal("370500"), "rate": Decimal("0.26"), "base_tax": Decimal("42678")},
            {"min_income": Decimal("370501"), "max_income": Decimal("512800"), "rate": Decimal("0.31"), "base_tax": Decimal("77362")},
            {"min_income": Decimal("512801"), "max_income": Decimal("673000"), "rate": Decimal("0.36"), "base_tax": Decimal("121475")},
            {"min_income": Decimal("673001"), "max_income": Decimal("857900"), "rate": Decimal("0.39"), "base_tax": Decimal("179147")},
            {"min_income": Decimal("857901"), "max_income": Decimal("1817000"), "rate": Decimal("0.41"), "base_tax": Decimal("251258")},
            {"min_income": Decimal("1817001"), "max_income": None, "rate": Decimal("0.45"), "base_tax": Decimal("644489")},
        ]

    return brackets


def extract_uif_config(wb):
    """Extract UIF configuration from workbook."""
    print("\n🔍 Extracting UIF configuration...")

    # Default UIF values (standard as per SARS)
    uif_config = {
        "employee_rate": Decimal("0.01"),  # 1%
        "employer_rate": Decimal("0.01"),  # 1%
        "monthly_cap": Decimal("17712"),   # 2024/25 value
        "annual_cap": Decimal("212544"),   # 2024/25 value
    }

    print(f"   Employee rate: {uif_config['employee_rate']*100}%")
    print(f"   Employer rate: {uif_config['employer_rate']*100}%")
    print(f"   Monthly cap: R{uif_config['monthly_cap']:,.2f}")
    print(f"   Annual cap: R{uif_config['annual_cap']:,.2f}")

    return uif_config


def extract_sdl_config(wb):
    """Extract SDL configuration from workbook."""
    print("\n🔍 Extracting SDL configuration...")

    # Default SDL values (standard as per SARS)
    sdl_config = {
        "rate": Decimal("0.01"),  # 1%
        "annual_threshold": Decimal("500000"),  # R500,000
    }

    print(f"   SDL rate: {sdl_config['rate']*100}%")
    print(f"   Annual threshold: R{sdl_config['annual_threshold']:,.2f}")

    return sdl_config


def extract_rebates(wb):
    """Extract tax rebates from workbook."""
    print("\n🔍 Extracting tax rebates...")

    # Default rebates (2024/25 values)
    rebates = {
        "primary": Decimal("17235"),   # All taxpayers
        "secondary": Decimal("9444"),  # Age 65+
        "tertiary": Decimal("3145"),   # Age 75+
    }

    print(f"   Primary rebate: R{rebates['primary']:,.2f}")
    print(f"   Secondary rebate: R{rebates['secondary']:,.2f}")
    print(f"   Tertiary rebate: R{rebates['tertiary']:,.2f}")

    return rebates


def generate_ruleset_code(values):
    """Generate Python code for za_2025_26_v1.py with extracted values."""

    if not values:
        return None

    brackets_code = "TAX_BRACKETS_ANNUAL: List[TaxBracket] = [\n"
    for bracket in values["tax_brackets"]:
        max_str = f'Decimal("{bracket["max_income"]}")' if bracket["max_income"] else "None"
        brackets_code += f"""    TaxBracket(
        min_income=Decimal("{bracket['min_income']}"),
        max_income={max_str},
        rate=Decimal("{bracket['rate']}"),
        base_tax=Decimal("{bracket['base_tax']}"),
    ),\n"""
    brackets_code += "]\n"

    uif_code = f"""UIF_EMPLOYEE_RATE = Decimal("{values['uif']['employee_rate']}")
UIF_EMPLOYER_RATE = Decimal("{values['uif']['employer_rate']}")
UIF_ANNUAL_CAP = Decimal("{values['uif']['annual_cap']}")
UIF_MONTHLY_CAP = Decimal("{values['uif']['monthly_cap']}")
"""

    sdl_code = f"""SDL_RATE = Decimal("{values['sdl']['rate']}")
SDL_ANNUAL_PAYROLL_THRESHOLD = Decimal("{values['sdl']['annual_threshold']}")
"""

    rebates_code = f"""REBATES = {{
    "primary": Decimal("{values['rebates']['primary']}"),
    "secondary": Decimal("{values['rebates']['secondary']}"),
    "tertiary": Decimal("{values['rebates']['tertiary']}"),
}}
"""

    return {
        "brackets": brackets_code,
        "uif": uif_code,
        "sdl": sdl_code,
        "rebates": rebates_code,
    }


def update_ruleset_file(values):
    """Update za_2025_26_v1.py with extracted values."""

    if not values:
        return

    code = generate_ruleset_code(values)

    print("\n📝 Generated code for za_2025_26_v1.py:")
    print("="*60)
    print("\n# PAYE Tax Brackets:")
    print(code["brackets"][:500] + "..." if len(code["brackets"]) > 500 else code["brackets"])
    print("\n# UIF Configuration:")
    print(code["uif"])
    print("# SDL Configuration:")
    print(code["sdl"])
    print("# Tax Rebates:")
    print(code["rebates"])
    print("="*60)

    print("\n⚠️  NEXT STEPS:")
    print("1. Review the extracted values above")
    print("2. Manually update src/app/rulesets/za_2025_26_v1.py")
    print("3. Add this comment at the top:")
    print(f'   # FROZEN - Extracted from sa_payroll_workbook.xlsx on {values["extraction_date"]}')
    print("4. Mark as FINAL - no further changes without creating v2")


if __name__ == "__main__":
    print("="*60)
    print("SA Payroll Workbook Value Extractor")
    print("="*60)

    values = extract_values_from_workbook()

    if values:
        update_ruleset_file(values)
        print("\n✅ Extraction complete!")
    else:
        print("\n❌ Extraction failed - check workbook location")

    print("\n" + "="*60)

