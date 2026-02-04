"""
Create test CSV and expected outputs from workbook for Task 7
"""

import openpyxl
from pathlib import Path
import csv
from decimal import Decimal

workbook_path = Path("data/samples/sa_payroll_workbook.xlsx")

print(f"Loading workbook from: {workbook_path}")

if not workbook_path.exists():
    print(f"ERROR: Workbook not found at {workbook_path}")
    exit(1)

wb = openpyxl.load_workbook(workbook_path, data_only=True)

print(f"Sheets: {wb.sheetnames}")

master_sheet = wb["Employee Master"]
outputs_sheet = wb["Outputs"]

# Create test CSV matching our canonical contract
test_csv_rows = []

# Header
header = [
    "payroll_run_id",
    "company_id",
    "pay_date",
    "tax_year",
    "payroll_frequency",
    "employee_id",
    "employment_type",
    "basic_salary",
    "overtime_pay",
    "bonus_commission",
    "allowances_taxable",
    "other_post_tax_deductions",
]

test_csv_rows.append(header)

# Get employee data from master
print("\nExtracting employee data from Employee Master:")
for row_idx in range(2, master_sheet.max_row + 1):
    emp_id = master_sheet.cell(row_idx, 3).value  # Column C
    basic_salary = master_sheet.cell(row_idx, 4).value  # Column D

    if not emp_id:
        print(f"  Row {row_idx}: No employee ID, stopping")
        break

    print(f"  {emp_id}: Basic R{basic_salary:,.2f if basic_salary else 0}")

    # For now, use basic salary only (workbook shows this is all we need)
    row_data = [
        "PAY_2025_03",  # payroll_run_id
        "TEST_COMPANY",  # company_id
        "2025-03-25",  # pay_date
        "2025_26",  # tax_year
        "monthly",  # payroll_frequency
        str(emp_id),  # employee_id
        "employee",  # employment_type
        str(basic_salary or 0),  # basic_salary
        "0",  # overtime_pay
        "0",  # bonus_commission
        "0",  # allowances_taxable
        "0",  # other_post_tax_deductions
    ]

    test_csv_rows.append(row_data)

# Save test CSV
csv_path = Path("data/samples/payroll_test_input.csv")
with open(csv_path, "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerows(test_csv_rows)

print(f"\n✅ Created test CSV: {csv_path}")
print(f"   Employees: {len(test_csv_rows) - 1}")

# Extract expected outputs
print("\nExtracting expected outputs from Outputs sheet:")
expected_outputs = []
totals = None

for row_idx in range(2, outputs_sheet.max_row + 1):
    emp_id = outputs_sheet.cell(row_idx, 1).value

    if not emp_id:
        print(f"  Row {row_idx}: No employee ID, stopping")
        break

    if str(emp_id).upper() == "TOTALS":
        # Extract totals
        totals = {
            "total_gross": Decimal(str(outputs_sheet.cell(row_idx, 2).value or 0)),
            "total_paye": Decimal(str(outputs_sheet.cell(row_idx, 3).value or 0)),
            "total_uif_employee": Decimal(str(outputs_sheet.cell(row_idx, 4).value or 0)),
            "total_uif_employer": Decimal(str(outputs_sheet.cell(row_idx, 5).value or 0)),
            "total_sdl": Decimal(str(outputs_sheet.cell(row_idx, 6).value or 0)),
            "total_net_pay": Decimal(str(outputs_sheet.cell(row_idx, 7).value or 0)),
            "total_employer_cost": Decimal(str(outputs_sheet.cell(row_idx, 8).value or 0)),
        }
        print(f"  TOTALS: Gross R{totals['total_gross']:,.2f}, PAYE R{totals['total_paye']:,.2f}")
        break

    expected = {
        "employee_id": str(emp_id),
        "gross_income": Decimal(str(outputs_sheet.cell(row_idx, 2).value or 0)),
        "paye": Decimal(str(outputs_sheet.cell(row_idx, 3).value or 0)),
        "uif_employee": Decimal(str(outputs_sheet.cell(row_idx, 4).value or 0)),
        "uif_employer": Decimal(str(outputs_sheet.cell(row_idx, 5).value or 0)),
        "sdl": Decimal(str(outputs_sheet.cell(row_idx, 6).value or 0)),
        "net_pay": Decimal(str(outputs_sheet.cell(row_idx, 7).value or 0)),
        "total_employer_cost": Decimal(str(outputs_sheet.cell(row_idx, 8).value or 0)),
    }
    expected_outputs.append(expected)
    print(f"  {emp_id}: PAYE R{expected['paye']:,.2f}, Net R{expected['net_pay']:,.2f}")

# Save expected outputs as JSON for test
import json
expected_data = {
    "employee_results": [
        {k: str(v) for k, v in result.items()}
        for result in expected_outputs
    ],
    "totals": {k: str(v) for k, v in totals.items()} if totals else None,
}

with open("data/samples/payroll_test_expected.json", "w") as f:
    json.dump(expected_data, f, indent=2)

print(f"\n✅ Created expected outputs: data/samples/payroll_test_expected.json")
print(f"   Test cases: {len(expected_outputs)}")
if totals:
    print(f"   Total PAYE: R{totals['total_paye']:,.2f}")
    print(f"   Total gross: R{totals['total_gross']:,.2f}")

print("\n✅ Test data creation complete!")

