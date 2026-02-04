"""
Generate expected output CSV from the Excel workbook (source of truth).

This script extracts calculated outputs from the workbook and writes them to CSV.
The end-to-end test uses this generated file instead of manually maintained expectations.

Usage:
    python scripts/generate_expected_from_workbook.py
    python scripts/generate_expected_from_workbook.py --workbook path/to/workbook.xlsx
"""

import argparse
import csv
import sys
from pathlib import Path
from decimal import Decimal

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def extract_outputs_from_workbook(workbook_path: Path) -> list[dict]:
    """
    Extract expected outputs from the workbook Outputs sheet.

    Args:
        workbook_path: Path to the Excel workbook.

    Returns:
        List of dicts with employee results.
    """
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    print(f"[*] Reading workbook: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    # Find the Outputs sheet
    if "Outputs" not in wb.sheetnames:
        raise ValueError(f"'Outputs' sheet not found. Available sheets: {wb.sheetnames}")

    outputs_sheet = wb["Outputs"]

    print(f"[*] Outputs sheet has {outputs_sheet.max_row} rows, {outputs_sheet.max_column} columns")

    # Read header to find column positions (find header row dynamically)
    headers = {}
    header_row = None
    for row_idx in range(1, min(outputs_sheet.max_row, 50) + 1):
        row_values = [outputs_sheet.cell(row_idx, c).value for c in range(1, outputs_sheet.max_column + 1)]
        if any(str(v).strip().lower() == "employee id" for v in row_values if v is not None):
            header_row = row_idx
            break

    if header_row is None:
        raise ValueError("Outputs header row not found (expected 'Employee ID').")

    for col_idx in range(1, outputs_sheet.max_column + 1):
        header = outputs_sheet.cell(header_row, col_idx).value
        if header:
            header_key = str(header).lower().strip().replace(" ", "_").replace("(", "").replace(")", "")
            headers[header_key] = col_idx
            print(f"   Column {col_idx}: {header} -> {header_key}")

    # Map to our expected field names
    col_map = {
        "employee_id": headers.get("employee_id"),
        "gross_income": headers.get("gross_pay_monthly", headers.get("gross_income")),
        "paye": headers.get("paye"),
        "uif_employee": headers.get("uif_employee"),
        "uif_employer": headers.get("uif_employer"),
        "sdl": headers.get("sdl"),
        "net_pay": headers.get("net_pay"),
        "total_employer_cost": headers.get("employer_cost", headers.get("total_employer_cost")),
    }

    missing_cols = [k for k, v in col_map.items() if v is None]
    if missing_cols:
        raise ValueError(f"Missing required columns in Outputs sheet: {missing_cols}")

    print(f"\n[*] Column mapping:")
    for field, col_idx in col_map.items():
        print(f"   {field} -> column {col_idx}")

    # Extract data
    expected_outputs = []

    print(f"\n[*] Extracting employee data...")

    for row_idx in range(header_row + 1, outputs_sheet.max_row + 1):
        emp_id = outputs_sheet.cell(row_idx, col_map["employee_id"]).value

        if not emp_id:
            print(f"   Row {row_idx}: Empty employee_id, stopping")
            break

        if str(emp_id).upper() == "TOTALS":
            print(f"   Row {row_idx}: TOTALS row, skipping")
            break

        result = {
            "employee_id": str(emp_id),
            "gross_income": _to_decimal(outputs_sheet.cell(row_idx, col_map["gross_income"]).value),
            "taxable_income": None,  # Not always in Outputs sheet
            "paye": _to_decimal(outputs_sheet.cell(row_idx, col_map["paye"]).value),
            "uif_employee": _to_decimal(outputs_sheet.cell(row_idx, col_map["uif_employee"]).value),
            "uif_employer": _to_decimal(outputs_sheet.cell(row_idx, col_map["uif_employer"]).value),
            "sdl": _to_decimal(outputs_sheet.cell(row_idx, col_map["sdl"]).value),
            "net_pay": _to_decimal(outputs_sheet.cell(row_idx, col_map["net_pay"]).value),
            "total_employer_cost": _to_decimal(outputs_sheet.cell(row_idx, col_map["total_employer_cost"]).value),
        }

        expected_outputs.append(result)
        print(f"   {emp_id}: Gross={result['gross_income']}, PAYE={result['paye']}, Net={result['net_pay']}")

    if not expected_outputs:
        raise ValueError("No employee data extracted from workbook Outputs sheet")

    print(f"\n[+] Extracted {len(expected_outputs)} employee results from workbook")

    return expected_outputs


def _to_decimal(value) -> Decimal:
    """Convert cell value to Decimal with 2-decimal precision."""
    if value is None or value == "":
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"))


def write_expected_csv(expected_outputs: list[dict], output_path: Path):
    """
    Write expected outputs to CSV.

    Args:
        expected_outputs: List of employee result dicts.
        output_path: Path to output CSV file.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "employee_id",
            "gross_income",
            "taxable_income",
            "paye",
            "uif_employee",
            "uif_employer",
            "sdl",
            "net_pay",
            "total_employer_cost",
        ])
        writer.writeheader()
        writer.writerows(expected_outputs)

    print(f"[+] Written expected outputs to: {output_path}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate expected output CSV from Excel workbook"
    )
    repo_root = Path(__file__).resolve().parents[1]
    parser.add_argument(
        "--workbook",
        type=Path,
        default=repo_root / "data" / "samples" / "sa_payroll_workbook.xlsx",
        help="Path to Excel workbook (default: data/samples/sa_payroll_workbook.xlsx)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/samples/payroll_expected_output_from_workbook.csv"),
        help="Output CSV path (default: data/samples/payroll_expected_output_from_workbook.csv)",
    )
    parser.add_argument(
        "--fallback",
        type=Path,
        default=None,
        help="Fallback CSV if workbook extraction fails",
    )

    args = parser.parse_args()

    if not args.workbook.is_absolute():
        args.workbook = (repo_root / args.workbook).resolve()

    print("="*60)
    print("Generate Expected Outputs from Workbook")
    print("="*60)
    print()

    try:
        # Extract from workbook
        expected_outputs = extract_outputs_from_workbook(args.workbook)

        # If no data extracted, try fallback
        if not expected_outputs and args.fallback and args.fallback.exists():
            print(f"\n[!] No data extracted from workbook")
            print(f"[*] Using fallback file: {args.fallback}")

            # Copy fallback to output
            import shutil
            shutil.copy(args.fallback, args.output)

            print(f"[+] Copied fallback to: {args.output}")
            print()
            print("="*60)
            print("[+] SUCCESS (using fallback)")
            print("="*60)
            print(f"Generated: {args.output}")
            print()
            print("Note: Using manually maintained fallback file.")
            print("To use workbook as source of truth, verify workbook has Outputs sheet with data.")

            return 0

        # Write to CSV
        write_expected_csv(expected_outputs, args.output)

        print()
        print("="*60)
        print("[+] SUCCESS")
        print("="*60)
        print(f"Generated: {args.output}")
        print(f"Employees: {len(expected_outputs)}")
        print()
        print("This file is now the source of truth for end-to-end tests.")

        return 0

    except Exception as e:
        print()
        print("="*60)
        print("[!] ERROR")
        print("="*60)
        print(f"{type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

        # Try fallback
        if args.fallback and args.fallback.exists():
            print()
            print("[!] Attempting to use fallback file...")
            try:
                import shutil
                shutil.copy(args.fallback, args.output)
                print(f"[+] Copied fallback to: {args.output}")
                return 0
            except Exception as e2:
                print(f"[!] Fallback also failed: {e2}")

        return 1


if __name__ == "__main__":
    sys.exit(main())

