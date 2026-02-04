"""
Create test CSV from the Inputs sheet
"""

import openpyxl
from pathlib import Path

workbook_path = Path("data/samples/sa_payroll_workbook.xlsx")
wb = openpyxl.load_workbook(workbook_path, data_only=True)

# Get employee master data
master_sheet = wb["Employee Master"]
inputs_sheet = wb["Inputs"]
outputs_sheet = wb["Outputs"]

print("Employee Master Data:")
for row_idx in range(2, master_sheet.max_row + 1):
    emp_id = master_sheet.cell(row_idx, 3).value
    if not emp_id:
        break
    name = master_sheet.cell(row_idx, 2).value
    basic = master_sheet.cell(row_idx, 4).value
    print(f"  {emp_id}: {name}, Basic: R{basic:,.2f if basic else 0}")

print("\nInputs Data:")
for row_idx in range(2, min(15, inputs_sheet.max_row + 1)):
    emp_id = inputs_sheet.cell(row_idx, 1).value
    if not emp_id:
        break
    basic = inputs_sheet.cell(row_idx, 2).value
    bonus = inputs_sheet.cell(row_idx, 3).value
    overtime = inputs_sheet.cell(row_idx, 4).value
    allowances = inputs_sheet.cell(row_idx, 5).value
    deductions = inputs_sheet.cell(row_idx, 6).value
    print(f"  {emp_id}: Basic={basic}, Bonus={bonus}, OT={overtime}, Allow={allowances}, Ded={deductions}")

print("\nOutputs Data:")
for row_idx in range(2, outputs_sheet.max_row + 1):
    emp_id = outputs_sheet.cell(row_idx, 1).value
    if not emp_id or emp_id == "TOTALS":
        if emp_id == "TOTALS":
            print(f"  TOTALS ROW:")
            for col_idx in range(2, 9):
                val = outputs_sheet.cell(row_idx, col_idx).value
                print(f"    Col {col_idx}: {val}")
        break
    gross = outputs_sheet.cell(row_idx, 2).value
    paye = outputs_sheet.cell(row_idx, 3).value
    uif_ee = outputs_sheet.cell(row_idx, 4).value
    uif_er = outputs_sheet.cell(row_idx, 5).value
    sdl = outputs_sheet.cell(row_idx, 6).value
    net = outputs_sheet.cell(row_idx, 7).value
    cost = outputs_sheet.cell(row_idx, 8).value
    print(f"  {emp_id}: Gross={gross}, PAYE={paye}, UIF_EE={uif_ee}, UIF_ER={uif_er}, SDL={sdl}, Net={net}, Cost={cost}")

