"""
Extract actual tax values and test data from sa_payroll_workbook.xlsx

This extracts:
1. PAYE tax brackets from "Ruleset v1.1 (2025-26)" sheet
2. UIF and SDL values
3. Test data from "Inputs" and "Outputs" sheets
"""

import openpyxl
from decimal import Decimal
from pathlib import Path
import csv
import io

workbook_path = Path("data/samples/sa_payroll_workbook.xlsx")

print("="*60)
print("EXTRACTING TAX VALUES AND TEST DATA")
print("="*60)

wb = openpyxl.load_workbook(workbook_path, data_only=True)

# Extract from Ruleset sheet
ruleset_sheet = wb["Ruleset v1.1 (2025-26)"]

print("\n📊 PAYE TAX BRACKETS (Annual):")
print("-" * 60)

brackets = []
# Start from row 5 (after headers)
for row_idx in range(5, 20):  # Check up to row 20
    from_val = ruleset_sheet.cell(row_idx, 1).value
    to_val = ruleset_sheet.cell(row_idx, 2).value
    base_tax = ruleset_sheet.cell(row_idx, 3).value
    rate = ruleset_sheet.cell(row_idx, 4).value

    if from_val and isinstance(from_val, (int, float)) and from_val > 0:
        bracket = {
            "min_income": Decimal(str(int(from_val))),
            "max_income": Decimal(str(int(to_val))) if to_val and to_val > 0 else None,
            "rate": Decimal(str(rate)) if rate else Decimal("0"),
            "base_tax": Decimal(str(int(base_tax))) if base_tax and base_tax > 0 else Decimal("0"),
        }
        brackets.append(bracket)
        max_str = f"R{bracket['max_income']:,}" if bracket['max_income'] else "∞"
        print(f"  R{bracket['min_income']:,} - {max_str}: {bracket['rate']*100}% (Base: R{bracket['base_tax']:,})")

print(f"\n✅ Extracted {len(brackets)} tax brackets")

# Extract rebates
print("\n📊 TAX REBATES (Annual):")
print("-" * 60)

rebates = {}
# Rebates are in columns 6-7, rows 5-7
primary = ruleset_sheet.cell(5, 7).value
secondary = ruleset_sheet.cell(6, 7).value
tertiary = ruleset_sheet.cell(7, 7).value

rebates = {
    "primary": Decimal(str(int(primary))) if primary else Decimal("17235"),
    "secondary": Decimal(str(int(secondary))) if secondary else Decimal("9444"),
    "tertiary": Decimal(str(int(tertiary))) if tertiary else Decimal("3145"),
}

print(f"  Primary: R{rebates['primary']:,}")
print(f"  Secondary (65+): R{rebates['secondary']:,}")
print(f"  Tertiary (75+): R{rebates['tertiary']:,}")

# Extract UIF and SDL
print("\n📊 UIF CONFIGURATION:")
print("-" * 60)

# UIF values are typically in rows 14-17
uif_employee_rate = Decimal("0.01")  # Default 1%
uif_employer_rate = Decimal("0.01")  # Default 1%
uif_monthly_cap = Decimal("17712")   # Default

# Try to extract from sheet
for row_idx in range(14, 20):
    label = ruleset_sheet.cell(row_idx, 1).value
    if label and "UIF" in str(label).upper():
        val = ruleset_sheet.cell(row_idx, 2).value
        if "Employee" in str(label):
            uif_employee_rate = Decimal(str(val)) if val else Decimal("0.01")
        elif "Employer" in str(label):
            uif_employer_rate = Decimal(str(val)) if val else Decimal("0.01")
        elif "Cap" in str(label) or "Ceiling" in str(label):
            uif_monthly_cap = Decimal(str(int(val))) if val else Decimal("17712")

print(f"  Employee rate: {uif_employee_rate * 100}%")
print(f"  Employer rate: {uif_employer_rate * 100}%")
print(f"  Monthly cap: R{uif_monthly_cap:,}")
print(f"  Annual cap: R{uif_monthly_cap * 12:,}")

print("\n📊 SDL CONFIGURATION:")
print("-" * 60)

sdl_rate = Decimal("0.01")  # Default 1%
sdl_threshold = Decimal("500000")  # Default

# Try to extract from sheet
for row_idx in range(18, 23):
    label = ruleset_sheet.cell(row_idx, 1).value
    if label and "SDL" in str(label).upper():
        val = ruleset_sheet.cell(row_idx, 2).value
        if "Rate" in str(label):
            sdl_rate = Decimal(str(val)) if val else Decimal("0.01")
        elif "Threshold" in str(label):
            sdl_threshold = Decimal(str(int(val))) if val else Decimal("500000")

print(f"  SDL rate: {sdl_rate * 100}%")
print(f"  Annual threshold: R{sdl_threshold:,}")

# Extract test data from Outputs sheet
print("\n📊 TEST DATA FROM OUTPUTS:")
print("-" * 60)

outputs_sheet = wb["Outputs"]

test_data = []
# Start from row 2 (after header)
for row_idx in range(2, outputs_sheet.max_row + 1):
    emp_id = outputs_sheet.cell(row_idx, 1).value
    if not emp_id or emp_id == "TOTALS":
        break

    test_case = {
        "employee_id": str(emp_id),
        "gross_pay": Decimal(str(outputs_sheet.cell(row_idx, 2).value or 0)),
        "paye": Decimal(str(outputs_sheet.cell(row_idx, 3).value or 0)),
        "uif_employee": Decimal(str(outputs_sheet.cell(row_idx, 4).value or 0)),
        "uif_employer": Decimal(str(outputs_sheet.cell(row_idx, 5).value or 0)),
        "sdl": Decimal(str(outputs_sheet.cell(row_idx, 6).value or 0)),
        "net_pay": Decimal(str(outputs_sheet.cell(row_idx, 7).value or 0)),
        "employer_cost": Decimal(str(outputs_sheet.cell(row_idx, 8).value or 0)),
    }
    test_data.append(test_case)
    print(f"  {emp_id}: Gross R{test_case['gross_pay']:,.2f}, PAYE R{test_case['paye']:,.2f}, Net R{test_case['net_pay']:,.2f}")

print(f"\n✅ Extracted {len(test_data)} test cases")

# Save extracted data
output = {
    "extraction_date": "2026-01-27",
    "source": "sa_payroll_workbook.xlsx",
    "tax_brackets": [
        {
            "min_income": str(b["min_income"]),
            "max_income": str(b["max_income"]) if b["max_income"] else None,
            "rate": str(b["rate"]),
            "base_tax": str(b["base_tax"]),
        }
        for b in brackets
    ],
    "rebates": {k: str(v) for k, v in rebates.items()},
    "uif": {
        "employee_rate": str(uif_employee_rate),
        "employer_rate": str(uif_employer_rate),
        "monthly_cap": str(uif_monthly_cap),
        "annual_cap": str(uif_monthly_cap * 12),
    },
    "sdl": {
        "rate": str(sdl_rate),
        "annual_threshold": str(sdl_threshold),
    },
    "test_cases": [
        {k: str(v) for k, v in tc.items()}
        for tc in test_data
    ],
}

import json
with open("data/samples/extracted_values.json", "w") as f:
    json.dump(output, f, indent=2)

print("\n✅ Saved extracted values to data/samples/extracted_values.json")
print("="*60)

