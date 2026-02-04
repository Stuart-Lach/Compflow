"""
Extract values and test data from sa_payroll_workbook.xlsx
"""

import openpyxl
from decimal import Decimal
from pathlib import Path
import json

workbook_path = Path("data/samples/sa_payroll_workbook.xlsx")

print("="*60)
print("EXTRACTING VALUES FROM WORKBOOK")
print("="*60)

wb = openpyxl.load_workbook(workbook_path, data_only=True)

print(f"\nSheets found: {wb.sheetnames}")

# Examine each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{sheet_name}:")
    print(f"  Max row: {sheet.max_row}, Max col: {sheet.max_column}")

    # Print first 10 rows to see structure
    print("  First 10 rows:")
    for row_idx in range(1, min(11, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(10, sheet.max_column + 1)):
            val = sheet.cell(row_idx, col_idx).value
            if val:
                row_data.append(f"{col_idx}:{val}")
        if row_data:
            print(f"    Row {row_idx}: {', '.join(row_data)}")

print("\n" + "="*60)

