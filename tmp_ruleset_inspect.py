from openpyxl import load_workbook
from pathlib import Path
wb_path = Path(r"C:\Users\adria\Compflow\compliance-engine\data\sa_payroll_workbook.xlsx")
print('exists', wb_path.exists())
wb = load_workbook(wb_path, data_only=True)
print('sheets', wb.sheetnames)
ws = wb["Ruleset v1.1 (2025-26)"]
for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=8, values_only=True):
    print(row)
